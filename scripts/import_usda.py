"""
Import foods from USDA FoodData Central + translate names with DeepL.
Outputs:
  - scripts/aliments_usda.sql     → import to Railway DB
  - scripts/aliments_usda.xlsx    → review with client

Setup:
  pip install requests deepl python-dotenv openpyxl

Usage:
  python scripts/import_usda.py
"""

import os
import sys
import time
import uuid
import requests
import deepl
from dotenv import load_dotenv

load_dotenv()

USDA_API_KEY  = os.environ.get("USDA_API_KEY", "")
DEEPL_API_KEY = os.environ.get("DEEPL_API_KEY", "")

if not USDA_API_KEY:
    sys.exit("❌  Falta USDA_API_KEY en .env")
if not DEEPL_API_KEY:
    sys.exit("❌  Falta DEEPL_API_KEY en .env")

MAX_FOODS  = 15_000
DATA_TYPES = ["Foundation", "SR Legacy"]
OUTPUT_SQL  = "scripts/aliments_usda.sql"
OUTPUT_XLSX = "scripts/aliments_usda.xlsx"

CATEGORY_MAP: dict[str, str] = {
    "Dairy and Egg Products":               "Lácteos y huevos",
    "Spices and Herbs":                     "Especias y hierbas",
    "Fats and Oils":                        "Grasas y aceites",
    "Poultry Products":                     "Aves",
    "Soups, Sauces, and Gravies":           "Salsas y sopas",
    "Sausages and Luncheon Meats":          "Embutidos",
    "Breakfast Cereals":                    "Cereales de desayuno",
    "Fruits and Fruit Juices":              "Frutas",
    "Pork Products":                        "Cerdo",
    "Vegetables and Vegetable Products":    "Verduras y vegetales",
    "Nut and Seed Products":                "Frutos secos y semillas",
    "Beef Products":                        "Res y vacuno",
    "Beverages":                            "Bebidas",
    "Finfish and Shellfish Products":       "Pescados y mariscos",
    "Legumes and Legume Products":          "Legumbres",
    "Lamb, Veal, and Game Products":        "Cordero y caza",
    "Baked Products":                       "Panadería y repostería",
    "Sweets":                               "Dulces",
    "Cereal Grains and Pasta":              "Granos y pastas",
    "Fast Foods":                           "Comida rápida",
    "Snacks":                               "Snacks",
    "Meals, Entrees, and Side Dishes":      "Platos preparados",
    "Restaurant Foods":                     "Restaurantes",
    "Ethnic Foods":                         "Cocina étnica",
    "American Indian/Alaska Native Foods":  "Otros",
    "Baby Foods":                           "Alimentos infantiles",
}

# Macronutrients
NUT_PROTEIN = 1003
NUT_FAT     = 1004
NUT_CARBS   = 1005
NUT_KCAL    = 1008

# Micronutrients → AlimentDescription columns
MICRONUTRIENTS = {
    "vita":               1106,  # Vitamin A (mcg)
    "vitb1":              1165,  # Thiamin B1 (mg)
    "vitb2":              1166,  # Riboflavin B2 (mg)
    "vitb3":              1167,  # Niacin B3 (mg)
    "vitb5":              1170,  # Pantothenic acid B5 (mg)
    "vitb6":              1175,  # Vitamin B6 (mg)
    "vitb9":              1177,  # Folate B9 (mcg)
    "vitb12":             1178,  # Vitamin B12 (mcg)
    "vitc":               1162,  # Vitamin C (mg)
    "vitd":               1114,  # Vitamin D (mcg)
    "vite":               1109,  # Vitamin E (mg)
    "vitk":               1185,  # Vitamin K (mcg)
    "calcium":            1087,  # Calcium (mg)
    "copper":             1098,  # Copper (mg)
    "iron":               1089,  # Iron (mg)
    "magnesium":          1090,  # Magnesium (mg)
    "manganese":          1101,  # Manganese (mg)
    "phosphorus":         1091,  # Phosphorus (mg)
    "potassium":          1092,  # Potassium (mg)
    "selenium":           1103,  # Selenium (mcg)
    "sodium":             1093,  # Sodium (mg)
    "zinc":               1095,  # Zinc (mg)
    "water":              1051,  # Water (g)
    "fiber":              1079,  # Fiber (g)
    "cholesterol":        1253,  # Cholesterol (mg)
    "saturated_fats":     1258,  # Saturated fat (g)
    "mono_saturated_fats":1292,  # Monounsaturated fat (g)
    "poli_saturated_fats":1293,  # Polyunsaturated fat (g)
}

EXCEL_HEADERS = [
    "Nombre (ES)", "Nombre (EN)", "Grupo", "Calorías (kcal)", "Proteínas (g)",
    "Carbohidratos (g)", "Grasas (g)", "Vitamina A (mcg)", "Vitamina B1 (mg)",
    "Vitamina B2 (mg)", "Vitamina B3 (mg)", "Vitamina B5 (mg)", "Vitamina B6 (mg)",
    "Vitamina B9 (mcg)", "Vitamina B12 (mcg)", "Vitamina C (mg)", "Vitamina D (mcg)",
    "Vitamina E (mg)", "Vitamina K (mcg)", "Calcio (mg)", "Cobre (mg)", "Hierro (mg)",
    "Magnesio (mg)", "Manganeso (mg)", "Fósforo (mg)", "Potasio (mg)", "Selenio (mcg)",
    "Sodio (mg)", "Zinc (mg)", "Agua (g)", "Fibra (g)", "Colesterol (mg)",
    "Grasas saturadas (g)", "Grasas monoinsaturadas (g)", "Grasas poliinsaturadas (g)",
    "USDA ID",
]


def get_nutrient(food: dict, nutrient_id: int) -> float | None:
    for n in food.get("foodNutrients", []):
        nid = n.get("nutrientId") or n.get("nutrient", {}).get("id")
        if nid == nutrient_id:
            val = n.get("value") or n.get("amount") or 0
            return round(float(val), 4)
    return None


def fetch_usda_page(page: int, page_size: int = 200) -> list[dict]:
    r = requests.get(
        "https://api.nal.usda.gov/fdc/v1/foods/search",
        params={
            "api_key":    USDA_API_KEY,
            "query":      "",
            "dataType":   DATA_TYPES,
            "pageSize":   page_size,
            "pageNumber": page,
        },
        timeout=30,
    )
    r.raise_for_status()
    return r.json().get("foods", [])


def translate_batch(translator: deepl.Translator, names: list[str]) -> list[str]:
    if not names:
        return []
    try:
        results = translator.translate_text(names, target_lang="ES", source_lang="EN")
        return [r.text for r in results]
    except Exception as e:
        print(f"  ⚠️  DeepL error: {e} — usando nombres en inglés")
        return names


def sql_escape(val) -> str:
    if val is None:
        return "NULL"
    if isinstance(val, str):
        return "'" + val.replace("\\", "\\\\").replace("'", "\\'") + "'"
    return str(val)


def main():
    translator = deepl.Translator(DEEPL_API_KEY)
    group_id_map: dict[str, int] = {}

    # ── Fetch ──────────────────────────────────────────────────────────────────
    print(f"🚀  Descargando hasta {MAX_FOODS:,} alimentos de USDA...\n")
    all_foods: list[dict] = []
    page = 1
    while len(all_foods) < MAX_FOODS:
        try:
            foods = fetch_usda_page(page)
        except Exception as e:
            print(f"  ⚠️  Error página {page}: {e}")
            break
        if not foods:
            break
        all_foods.extend(foods)
        print(f"   Página {page}: {len(all_foods):,} alimentos")
        page += 1
        time.sleep(0.3)

    all_foods = all_foods[:MAX_FOODS]
    print(f"\n✅  {len(all_foods):,} alimentos descargados\n")

    for food in all_foods:
        cat_en = food.get("foodCategory") or food.get("foodCategoryLabel") or ""
        group_name = CATEGORY_MAP.get(cat_en, "Otros")
        if group_name not in group_id_map:
            group_id_map[group_name] = len(group_id_map) + 1

    # ── Translate ──────────────────────────────────────────────────────────────
    print("🌐  Traduciendo al español...\n")
    en_names = [f.get("description", "Unknown") for f in all_foods]
    es_names: list[str] = []
    batch_size = 50
    for i in range(0, len(en_names), batch_size):
        batch = en_names[i:i + batch_size]
        es_names.extend(translate_batch(translator, batch))
        print(f"   {min(i + batch_size, len(en_names)):,} / {len(en_names):,}")
        time.sleep(0.1)

    # ── Build rows ─────────────────────────────────────────────────────────────
    rows = []
    for food, es_name in zip(all_foods, es_names):
        cat_en     = food.get("foodCategory") or food.get("foodCategoryLabel") or ""
        group_name = CATEGORY_MAP.get(cat_en, "Otros")
        fdc_id     = food["fdcId"]

        macros = {
            "proteins":      get_nutrient(food, NUT_PROTEIN),
            "carbohydrates": get_nutrient(food, NUT_CARBS),
            "fats":          get_nutrient(food, NUT_FAT),
            "calories":      get_nutrient(food, NUT_KCAL),
        }
        micros = {col: get_nutrient(food, nid) for col, nid in MICRONUTRIENTS.items()}

        rows.append({
            "id":         str(uuid.uuid4()),
            "fdc_id":     fdc_id,
            "es_name":    es_name[:255],
            "en_name":    food.get("description", "")[:255],
            "group_name": group_name,
            **macros,
            **micros,
        })

    # ── Write SQL ──────────────────────────────────────────────────────────────
    print(f"\n📝  Generando {OUTPUT_SQL}...")
    with open(OUTPUT_SQL, "w", encoding="utf-8") as f:
        f.write("-- USDA FoodData Central import con micronutrientes\n")
        f.write("SET NAMES utf8mb4;\n\n")

        f.write("-- Grupos\n")
        for name in group_id_map:
            f.write(
                f"INSERT INTO group_foods (name, status, created_at, updated_at) "
                f"VALUES ({sql_escape(name)}, 1, NOW(), NOW()) "
                f"ON DUPLICATE KEY UPDATE name=name;\n"
            )
        f.write("\n-- Alimentos\n")

        micro_cols = ", ".join(MICRONUTRIENTS.keys())
        micro_vals_tpl = ", ".join(["{" + k + "}" for k in MICRONUTRIENTS.keys()])

        for row in rows:
            fdc_comment = f"usda:{row['fdc_id']}"
            micro_vals  = ", ".join(sql_escape(row[k]) for k in MICRONUTRIENTS.keys())
            aid = row["id"]

            f.write(
                f"INSERT INTO aliments "
                f"(id, name, group_food_id, proteins, carbohydrates, fats, calories, "
                f"quantity, comments, created_at, updated_at) "
                f"SELECT {sql_escape(aid)}, {sql_escape(row['es_name'])}, "
                f"(SELECT id FROM group_foods WHERE name={sql_escape(row['group_name'])} LIMIT 1), "
                f"{sql_escape(row['proteins'])}, {sql_escape(row['carbohydrates'])}, "
                f"{sql_escape(row['fats'])}, {sql_escape(row['calories'])}, "
                f"100, {sql_escape(fdc_comment)}, NOW(), NOW() "
                f"WHERE NOT EXISTS (SELECT 1 FROM aliments WHERE comments={sql_escape(fdc_comment)});\n"
            )
            f.write(
                f"INSERT INTO aliment_descriptions "
                f"(aliment_id, {micro_cols}) "
                f"SELECT id, {micro_vals} FROM aliments WHERE comments={sql_escape(fdc_comment)} "
                f"AND NOT EXISTS (SELECT 1 FROM aliment_descriptions ad "
                f"JOIN aliments a ON ad.aliment_id=a.id WHERE a.comments={sql_escape(fdc_comment)});\n"
            )

    # ── Write Excel ────────────────────────────────────────────────────────────
    print(f"📊  Generando {OUTPUT_XLSX}...")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Alimentos USDA"

        header_fill = PatternFill("solid", fgColor="5B2D8E")
        header_font = Font(bold=True, color="FFFFFF", size=10)

        for col_idx, header in enumerate(EXCEL_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for r_idx, row in enumerate(rows, 2):
            micro_values = [row.get(k) for k in MICRONUTRIENTS.keys()]
            ws.append([
                row["es_name"], row["en_name"], row["group_name"],
                row["calories"], row["proteins"], row["carbohydrates"], row["fats"],
                *micro_values,
                row["fdc_id"],
            ])

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 30)

        ws.freeze_panes = "A2"
        wb.save(OUTPUT_XLSX)
        print(f"   ✅  Excel guardado: {OUTPUT_XLSX}")
    except ImportError:
        print("   ⚠️  openpyxl no instalado — ejecuta: pip install openpyxl")

    print(f"\n🎉  Listo: {len(rows):,} alimentos")
    print(f"    SQL  → {OUTPUT_SQL}")
    print(f"    Excel → {OUTPUT_XLSX}")
    print("\n── Importar a Railway ──────────────────────────────────────────")
    print("  Get-Content scripts\\aliments_usda.sql | railway run mysql railway")
    print("────────────────────────────────────────────────────────────────")


if __name__ == "__main__":
    main()
